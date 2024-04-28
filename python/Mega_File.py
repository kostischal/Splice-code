#! python3

import openpyxl as xl
import os, re, send2trash
import pandas as pd
from openpyxl.styles import Font,Color,Border
from copy import copy

os.chdir('C:\\Users\\user\\Desktop')

file1='Thread_Data.xlsx'#first file
wb1=xl.load_workbook(file1)
ws1=wb1.worksheets[0]
ws2=wb1.create_sheet()

mr=ws1.max_row

for i in range(9,mr+1):
    c=ws1.cell(row=i,column=4)
    match=re.search(r'^2FO|8FO|12FO',str(c.value))#values which start with,from thread export to be sace in daybook
    if match:
        ws2.cell(row=i-10,column=1).value=c.value
    else:
        continue


mr=ws2.max_row

for i in range(1,mr+1):#delete unnecessary rows in sheet_data
   c=ws2.cell(row=i,column=1) 
   if c.value is None:
        ws2.delete_rows(idx=i)
   else:
        continue

for i in range(1,mr+1,2):#insert the same value on next row below in sheet_data
    c=ws2.cell(row=i,column=1)
    ws2.cell(row=i+1,column=1).value=c.value

for i in range(1,mr+2):#format the cell values as we want for the final sc_template
    c=ws2.cell(row=i,column=1)
    if c.value is not None:
        ws2.cell(row=i,column=1).value=str(c.value).replace(' - ','-')
        ws2.cell(row=i,column=1).value=str(c.value).replace('m','')
        ws2.cell(row=i,column=1).value=str(c.value).replace('.',',')
    else:
        break


ws2.cell(row=1,column=2).value=str(ws1.cell(row=9,column=4).value).replace(' - ','-')#format the value about name of feeder from thread



wb1.save(str('Thread_Data.xlsx'))


file2='Address.xlsx'#second file
wb2=xl.load_workbook(file2)
ws=wb2.worksheets[0]

mr=ws.max_row

for i in range(9,mr+1):#format the cell values as we want for the final sc_template
    c2=ws.cell(row=i,column=3)
    ws.cell(row=i,column=3).value=c2.value.replace(' ','')

wb2.save(str('Address.xlsx'))


file1='Thread_Data.xlsx'#first file load
wb1=xl.load_workbook(file1)
ws1=wb1.worksheets[2]

file2='Address.xlsx'#second file load
wb2=xl.load_workbook(file2)
ws2=wb2.worksheets[0]

file3='sc_template.xlsx'#sc_template file load and active
wb3=xl.load_workbook(file3)
ws3=wb3.active

mr1=ws1.max_row
mr2=ws2.max_row
mr3=ws3.max_row

x=1
a=1
b=1


for i in range(1,mr1+1):#fill columns M(cable),U(cable),V(cable length),X(label)
    c=ws1.cell(row=i,column=1)
    if c.value is not None:
        ws3.cell(row=i+3,column=13).value=str(c.value).split(' ')[2]
        ws3.cell(row=i+3,column=21).value=str(c.value).split(' ')[2]
        ws3.cell(row=i+3,column=22).value=str(c.value).split(' ')[3]
        ws3.cell(row=i+3,column=24).value=str(c.value).split(' ')[2].split('-')[1]
    else:
        break

ws3.cell(row=1,column=2).value=str(ws1.cell(row=1,column=1).value).split(' ')[2].split('-')[0]#fill 'B1' column of SC
ws3.cell(row=4,column=5).value=str(ws1.cell(row=1,column=2).value).split(' ')[3]#fill 'E4' column of SC

for i in range(9,mr2 +1):#fill the column address based on key value 'label'
    c1=ws2.cell(row=i,column=2)
    c2=ws2.cell(row=i,column=3)
    c3=ws2.cell(row=i,column=6)
    for i in range(1,mr3 +1):
        if c3.value==str(ws3.cell(row=i+3,column=24).value).split('_')[0]:
            ws3.cell(row=i+3,column=23).value=' '.join([str(c1.value),str(c2.value)])
        else:
            continue



for i in range(4,mr3+1):#fill the second column 'fiber'
     c1=ws3.cell(row=i,column=13)
     c2=ws3.cell(row=i+1,column=13)          
     if (c1.value and c2.value) is not None:

         if c1.value==c2.value:
            ws3.cell(row=i,column=11).value=x
            ws3.cell(row=i+1,column=11).value=x+1
            x=x+1
         else:
            ws3.cell(row=i+1,column=11).value=1
            x=1
     else:
         break

for i in range(4,mr3+1):#fill the first column 'fiber'
    c1=ws3.cell(row=i,column=13)
    if c1.value is not None:
        ws3.cell(row=i,column=9).value=a
        a=a+1
    else:
        break

for i in range(4,mr3+1,2):#fill the column 'cassette'
    c1=ws3.cell(row=i,column=13)
    if c1.value is not None:
        ws3.cell(row=i,column=10).value=b
        ws3.cell(row=i+1,column=10).value=b
        b=b+1
    else:
        break

wb3.save(str('sc_template2.xlsx'))#final file
#send2trash.send2trash(file1)
#send2trash.send2trash(file2)

excel_file='sc_template2.xlsx'#pandas

df=pd.read_excel(excel_file,usecols=[12],skiprows=[1,2])
df2=df.value_counts()

df2.to_excel('help_file.xlsx')#help_file with count of cable values 

file1='help_file.xlsx'
wb1=xl.load_workbook(file1)
ws1=wb1.worksheets[0]

file2='sc_template2.xlsx'
wb2=xl.load_workbook(file2)
ws2=wb2.worksheets[0]
ws3=wb2.worksheets[1]

mr1=ws1.max_row
mr2=ws2.max_row

for i in range(2,mr1+1):#fill the column 'capacity' based on values of help_file
    c1=ws1.cell(row=i,column=1)
    c2=ws1.cell(row=i,column=2)
    for i in range(4,mr2+1):
        if c1.value==ws2.cell(row=i,column=13).value:
            if c2.value==2:
                ws2.cell(row=i,column=12).value=ws3.cell(row=21,column=6).value
                if ws3.cell(row=21,column=6).has_style:
                    ws2.cell(row=i,column=12)._style=copy(ws3.cell(row=21,column=6)._style)
                #ws2.cell(row=i,column=12).font=Font(bold=True,color="0033CCCC")
            elif c2.value>2 and c2.value<=8:
                ws2.cell(row=i,column=12).value=ws3.cell(row=20,column=6).value
                if ws3.cell(row=20,column=6).has_style:
                    ws2.cell(row=i,column=12)._style=copy(ws3.cell(row=20,column=6)._style)
                #ws2.cell(row=i,column=12).font=Font(bold=True,color="00883300")
            else:
                ws2.cell(row=i,column=12).value=ws3.cell(row=19,column=6).value
                if ws3.cell(row=19,column=6).has_style:
                    ws2.cell(row=i,column=12)._style=copy(ws3.cell(row=19,column=6)._style)
                #ws2.cell(row=i,column=12).font=Font(bold=True,color="00993366")
        else:
            continue


       
wb2.save(str('sc_template2.xlsx'))
send2trash.send2trash(file1)

